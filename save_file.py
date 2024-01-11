import openpyxl

class Excel_Save():

    def __init__(self, save):
        self.save = save
        self.workbook = self.create_or_load_excel()

    def create_or_load_excel(self):
        try:
            workbook = openpyxl.load_workbook(self.save)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            
            workbook.active.title = "Data"
            
            sheet = workbook.active
            sheet["A1"] = "İD"
            sheet["B1"] = "Name"
            sheet["C1"] = "Surname"
            sheet["D1"] = "E-mail"
            sheet["E1"] = "Date_of_Employment"
            sheet["F1"] = "Salary"
            sheet["G1"] = "Department"
            sheet["H1"] = "Language"
        
        return workbook
    
    def add_data(self, data):
        sheet = self.workbook.active

        for row in data:
            sheet.append(row)
        
        self.workbook.save(self.save)
    
    def get_row_by_id(self, target_id):
        sheet = self.workbook.active
     
        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
            if row[0] == target_id:  
                return row
        return None  
    
    def get_all_data(self):
        sheet = self.workbook.active
        all_data = []

        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
            all_data.append(row)

        return all_data
    
    def update_salary_by_id(self, target_id, new_salary):
        sheet = self.workbook.active
        found_row = None

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True), start=2):
            if row[0] == target_id:
                found_row = row_index
                break

        if found_row is not None:
            sheet.cell(row=found_row, column=6, value=new_salary)  
            self.workbook.save(self.save)
    
    def delete_row_by_id(self, target_id):
        sheet = self.workbook.active
        found_row = None

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True), start=2):
            if row[0] == target_id:
                found_row = row_index
                break

        if found_row is not None:
            sheet.delete_rows(found_row)
            self.workbook.save(self.save)


